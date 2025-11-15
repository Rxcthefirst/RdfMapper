#!/usr/bin/env python3
"""Create test Excel file with multiple sheets for testing."""

import sys
sys.path.insert(0, 'src')

import polars as pl
from pathlib import Path

print("="*80)
print("Creating Multi-Sheet Test Data")
print("="*80)

# Create output directory
output_dir = Path("test_data/multisheet")
output_dir.mkdir(parents=True, exist_ok=True)

# Create sample data
print("\n1. Creating Orders sheet...")
orders_data = {
    "OrderID": ["ORD001", "ORD002", "ORD003", "ORD004", "ORD005"],
    "CustomerID": ["CUST001", "CUST002", "CUST001", "CUST003", "CUST002"],
    "OrderDate": ["2024-01-15", "2024-01-16", "2024-01-17", "2024-01-18", "2024-01-19"],
    "TotalAmount": [150.00, 89.99, 220.50, 45.00, 175.00],
    "Status": ["Completed", "Shipped", "Completed", "Processing", "Shipped"]
}
orders_df = pl.DataFrame(orders_data)
print(f"  ✓ Created {len(orders_df)} orders")

print("\n2. Creating Customers sheet...")
customers_data = {
    "CustomerID": ["CUST001", "CUST002", "CUST003"],
    "FirstName": ["John", "Jane", "Bob"],
    "LastName": ["Doe", "Smith", "Johnson"],
    "Email": ["john@example.com", "jane@example.com", "bob@example.com"],
    "Phone": ["555-0101", "555-0102", "555-0103"],
    "City": ["New York", "Los Angeles", "Chicago"]
}
customers_df = pl.DataFrame(customers_data)
print(f"  ✓ Created {len(customers_df)} customers")

print("\n3. Creating OrderItems sheet...")
order_items_data = {
    "ItemID": ["ITEM001", "ITEM002", "ITEM003", "ITEM004", "ITEM005", "ITEM006"],
    "OrderID": ["ORD001", "ORD001", "ORD002", "ORD003", "ORD003", "ORD004"],
    "ProductID": ["PROD001", "PROD002", "PROD001", "PROD003", "PROD002", "PROD001"],
    "Quantity": [2, 1, 3, 1, 2, 1],
    "Price": [50.00, 50.00, 29.99, 150.00, 70.50, 45.00]
}
order_items_df = pl.DataFrame(order_items_data)
print(f"  ✓ Created {len(order_items_df)} order items")

print("\n4. Creating Products sheet...")
products_data = {
    "ProductID": ["PROD001", "PROD002", "PROD003"],
    "ProductName": ["Widget A", "Widget B", "Widget C"],
    "Category": ["Electronics", "Electronics", "Furniture"],
    "Price": [50.00, 29.99, 150.00],
    "Stock": [100, 50, 25]
}
products_df = pl.DataFrame(products_data)
print(f"  ✓ Created {len(products_df)} products")

# Write to Excel with multiple sheets
output_file = output_dir / "ecommerce_orders.xlsx"

print(f"\n5. Writing to Excel file: {output_file}...")
try:
    # Polars doesn't support multi-sheet writing directly, use openpyxl
    from openpyxl import Workbook

    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Add Orders sheet
    ws_orders = wb.create_sheet("Orders")
    ws_orders.append(list(orders_data.keys()))
    for row in zip(*orders_data.values()):
        ws_orders.append(list(row))

    # Add Customers sheet
    ws_customers = wb.create_sheet("Customers")
    ws_customers.append(list(customers_data.keys()))
    for row in zip(*customers_data.values()):
        ws_customers.append(list(row))

    # Add OrderItems sheet
    ws_items = wb.create_sheet("OrderItems")
    ws_items.append(list(order_items_data.keys()))
    for row in zip(*order_items_data.values()):
        ws_items.append(list(row))

    # Add Products sheet
    ws_products = wb.create_sheet("Products")
    ws_products.append(list(products_data.keys()))
    for row in zip(*products_data.values()):
        ws_products.append(list(row))

    # Save workbook
    wb.save(output_file)

    print("  ✓ Excel file created successfully")
except ImportError:
    print("  ✗ Error: openpyxl not installed")
    print("  Please install: pip install openpyxl")
    sys.exit(1)
except Exception as e:
    print(f"  ✗ Error: {e}")
    import traceback
    traceback.print_exc()
    sys.exit(1)

# Create relationship diagram
print("\n6. Data relationships:")
print("  Orders.CustomerID → Customers.CustomerID (many-to-one)")
print("  OrderItems.OrderID → Orders.OrderID (many-to-one)")
print("  OrderItems.ProductID → Products.ProductID (many-to-one)")

print("\n" + "="*80)
print("Multi-Sheet Test Data Created Successfully!")
print("="*80)
print(f"\nFile: {output_file}")
print("\nSheets:")
print("  1. Orders (5 rows)")
print("  2. Customers (3 rows)")
print("  3. OrderItems (6 rows)")
print("  4. Products (3 rows)")
print("\nTo test:")
print(f"  rdfmap generate \\")
print(f"    --ontology examples/mortgage/ontology/mortgage.ttl \\")
print(f"    --data {output_file} \\")
print(f"    --output test_multisheet_mapping.yaml")


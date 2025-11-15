#!/usr/bin/env python
"""Quick test for relationship detection fix."""

from pathlib import Path
from openpyxl import Workbook
from rdfmap.generator.multisheet_analyzer import MultiSheetAnalyzer

# Create test file
tmp_path = Path("./test_fix.xlsx")
wb = Workbook()
wb.remove(wb.active)

# Create Customers sheet
ws_customers = wb.create_sheet("Customers")
ws_customers.append(["CustomerID", "Name", "Email"])
ws_customers.append(["C001", "John Doe", "john@example.com"])
ws_customers.append(["C002", "Jane Smith", "jane@example.com"])

# Create Orders sheet
ws_orders = wb.create_sheet("Orders")
ws_orders.append(["OrderID", "CustomerID", "OrderDate", "Total"])
ws_orders.append(["O001", "C001", "2024-01-15", 100.00])
ws_orders.append(["O002", "C002", "2024-01-16", 200.00])
ws_orders.append(["O003", "C001", "2024-01-17", 150.00])

wb.save(tmp_path)

# Test
analyzer = MultiSheetAnalyzer(str(tmp_path))

# Test _find_matching_sheet directly
print("Testing _find_matching_sheet:")
print(f"  'Customer' -> {analyzer._find_matching_sheet('Customer')}")
print(f"  'Order' -> {analyzer._find_matching_sheet('Order')}")

relationships = analyzer.detect_relationships()
print(f"\nRelationships: {len(relationships)}")

tmp_path.unlink()

print(f"✓ Test {'PASSED' if len(relationships) > 0 else 'FAILED'}")


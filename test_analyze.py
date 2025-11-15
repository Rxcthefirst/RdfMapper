#!/usr/bin/env python
"""Test _analyze_relationship in isolation."""
from pathlib import Path
from openpyxl import Workbook
from rdfmap.generator.multisheet_analyzer import MultiSheetAnalyzer
import traceback

tmp_path = Path("./test_analyze.xlsx")
wb = Workbook()
wb.remove(wb.active)

ws_customers = wb.create_sheet("Customers")
ws_customers.append(["CustomerID", "Name", "Email"])
ws_customers.append(["C001", "John Doe", "john@example.com"])
ws_customers.append(["C002", "Jane Smith", "jane@example.com"])

ws_orders = wb.create_sheet("Orders")
ws_orders.append(["OrderID", "CustomerID", "OrderDate", "Total"])
ws_orders.append(["O001", "C001", "2024-01-15", 100.00])
ws_orders.append(["O002", "C002", "2024-01-16", 200.00])
ws_orders.append(["O003", "C001", "2024-01-17", 150.00])

wb.save(tmp_path)

analyzer = MultiSheetAnalyzer(str(tmp_path))

# Get the sheets
orders_info = analyzer.sheets["Orders"]
customers_info = analyzer.sheets["Customers"]

print("Orders info:")
print(f"  sample_data: {orders_info.sample_data}")
print(f"  columns: {orders_info.column_names}")

print("\nCustomers info:")
print(f"  sample_data: {customers_info.sample_data}")
print(f"  columns: {customers_info.column_names}")

# Try to call _analyze_relationship
print("\n" + "="*60)
print("Calling _analyze_relationship")
print("="*60)

try:
    result = analyzer._analyze_relationship(
        source_sheet="Orders",
        target_sheet="Customers",
        source_col="CustomerID",
        target_col="CustomerID",
        source_info=orders_info,
        target_info=customers_info
    )
    print(f"\nResult: {result}")

    if result:
        print(f"  ✓ Relationship created!")
        print(f"    Type: {result.relationship_type}")
        print(f"    Confidence: {result.confidence}")
    else:
        print(f"  ❌ Result is None")

except Exception as e:
    print(f"\n❌ Exception occurred: {e}")
    traceback.print_exc()

tmp_path.unlink()


#!/usr/bin/env python
"""Direct inline test to understand the issue."""
import sys
import warnings

# Capture warnings
warnings.simplefilter("always")

from pathlib import Path
from openpyxl import Workbook

tmp_path = Path("./test_inline.xlsx")
wb = Workbook()
wb.remove(wb.active)

ws_customers = wb.create_sheet("Customers")
ws_customers.append(["CustomerID", "Name", "Email"])
ws_customers.append(["C001", "John Doe", "john@example.com"])
ws_customers.append(["C002", "Jane Smith", "jane@example.com"])

ws_orders = wb.create_sheet("Orders")
ws_orders.append(["OrderID", "CustomerID", "OrderDate", "Total"])
ws_orders.append(["O001", "C001", "2024-01-15", 100.00])
ws_orders.append(["O002", "C002", "2024-01-16", 200.00])
ws_orders.append(["O003", "C001", "2024-01-17", 150.00])

wb.save(tmp_path)

print("Testing multisheet analyzer...")

from rdfmap.generator.multisheet_analyzer import MultiSheetAnalyzer

analyzer = MultiSheetAnalyzer(str(tmp_path))

print(f"Sheets loaded: {list(analyzer.sheets.keys())}")

# Check Orders sheet
orders = analyzer.sheets["Orders"]
print(f"\nOrders FK candidates: {orders.foreign_key_candidates}")

# Check if we can find matching sheet
if "CustomerID" in orders.foreign_key_candidates:
    entity = orders.foreign_key_candidates["CustomerID"]
    print(f"Entity extracted from CustomerID: '{entity}'")
    match = analyzer._find_matching_sheet(entity)
    print(f"Matching sheet: {match}")

# Now detect relationships
print("\n" + "="*60)
relationships = analyzer.detect_relationships()
print(f"Relationships detected: {len(relationships)}")

for rel in relationships:
    print(f"  {rel.source_sheet}.{rel.source_column} -> {rel.target_sheet}.{rel.target_column}")
    print(f"  Type: {rel.relationship_type}, Confidence: {rel.confidence:.2f}")

tmp_path.unlink()

if len(relationships) > 0:
    print("\n✓ SUCCESS")
    sys.exit(0)
else:
    print("\n❌ FAILED - No relationships detected")
    sys.exit(1)


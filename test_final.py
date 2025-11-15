#!/usr/bin/env python
"""Final comprehensive test."""
import sys
from pathlib import Path
from openpyxl import Workbook

print("="*70)
print("COMPREHENSIVE MULTISHEET ANALYZER TEST")
print("="*70)

tmp_path = Path("./test_final.xlsx")

# Create test Excel file
wb = Workbook()
wb.remove(wb.active)

ws_customers = wb.create_sheet("Customers")
ws_customers.append(["CustomerID", "Name", "Email"])
ws_customers.append(["C001", "John Doe", "john@example.com"])
ws_customers.append(["C002", "Jane Smith", "jane@example.com"])

ws_orders = wb.create_sheet("Orders")
ws_orders.append(["OrderID", "CustomerID", "OrderDate", "Total"])
ws_orders.append(["O001", "C001", "2024-01-15", 100.00])
ws_orders.append(["O002", "C002", "2024-01-16", 200.00])
ws_orders.append(["O003", "C001", "2024-01-17", 150.00])

wb.save(tmp_path)
print("\n✓ Test Excel file created")

try:
    from rdfmap.generator.multisheet_analyzer import MultiSheetAnalyzer

    # Initialize analyzer
    analyzer = MultiSheetAnalyzer(str(tmp_path))
    print(f"✓ Analyzer initialized")
    print(f"  Sheets: {list(analyzer.sheets.keys())}")

    # Check Orders sheet analysis
    orders = analyzer.sheets["Orders"]
    print(f"\n✓ Orders sheet analyzed")
    print(f"  Identifier columns: {orders.identifier_columns}")
    print(f"  FK candidates: {orders.foreign_key_candidates}")

    # Check Customers sheet analysis
    customers = analyzer.sheets["Customers"]
    print(f"\n✓ Customers sheet analyzed")
    print(f"  Identifier columns: {customers.identifier_columns}")
    print(f"  FK candidates: {customers.foreign_key_candidates}")

    # Test _find_matching_sheet
    if "CustomerID" in orders.foreign_key_candidates:
        entity = orders.foreign_key_candidates["CustomerID"]
        match = analyzer._find_matching_sheet(entity)
        print(f"\n✓ _find_matching_sheet('{entity}') = {match}")

        if match != "Customers":
            print(f"  ❌ Expected 'Customers', got '{match}'")
            sys.exit(1)

    # Test _find_primary_key
    if match:
        pk = analyzer._find_primary_key(customers, entity)
        print(f"✓ _find_primary_key(..., '{entity}') = {pk}")

        if pk != "CustomerID":
            print(f"  ❌ Expected 'CustomerID', got '{pk}'")
            sys.exit(1)

    # Test detect_relationships
    print(f"\n{'='*70}")
    print("TESTING detect_relationships()")
    print("="*70)

    relationships = analyzer.detect_relationships()

    print(f"\n✓ detect_relationships() completed")
    print(f"  Found {len(relationships)} relationship(s)")

    if len(relationships) == 0:
        print("\n  ❌ ERROR: No relationships detected!")
        print("\n  This is unexpected. The relationship should have been found.")
        print("  Debugging information:")
        print(f"    - Orders.CustomerID exists: {'CustomerID' in orders.column_names}")
        print(f"    - Customers.CustomerID exists: {'CustomerID' in customers.column_names}")
        print(f"    - Orders sample data: {orders.sample_data}")
        print(f"    - Customers sample data: {customers.sample_data}")
        sys.exit(1)

    # Validate the relationship
    for i, rel in enumerate(relationships, 1):
        print(f"\n  Relationship {i}:")
        print(f"    Source: {rel.source_sheet}.{rel.source_column}")
        print(f"    Target: {rel.target_sheet}.{rel.target_column}")
        print(f"    Type: {rel.relationship_type}")
        print(f"    Confidence: {rel.confidence:.2f}")
        if rel.cardinality_analysis:
            print(f"    Cardinality analysis:")
            for key, value in rel.cardinality_analysis.items():
                print(f"      - {key}: {value}")

    # Check if we found the expected relationship
    found_expected = False
    for rel in relationships:
        if (rel.source_sheet == "Orders" and
            rel.target_sheet == "Customers" and
            "CustomerID" in rel.source_column and
            "CustomerID" in rel.target_column):
            found_expected = True
            break

    if found_expected:
        print("\n" + "="*70)
        print("✓ TEST PASSED")
        print("="*70)
        print("\nThe expected relationship was found:")
        print("  Orders.CustomerID -> Customers.CustomerID")
        sys.exit(0)
    else:
        print("\n" + "="*70)
        print("❌ TEST FAILED")
        print("="*70)
        print("\nExpected relationship not found:")
        print("  Expected: Orders.CustomerID -> Customers.CustomerID")
        sys.exit(1)

except Exception as e:
    print(f"\n❌ Exception occurred: {e}")
    import traceback
    traceback.print_exc()
    sys.exit(1)

finally:
    if tmp_path.exists():
        tmp_path.unlink()
        print("\n✓ Test file cleaned up")


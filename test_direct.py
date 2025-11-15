#!/usr/bin/env python
"""Direct test of multisheet analyzer."""
import sys
from pathlib import Path
from openpyxl import Workbook

# Create exact same test file as in test
tmp_path = Path("./test_direct.xlsx")
wb = Workbook()
wb.remove(wb.active)

ws_customers = wb.create_sheet("Customers")
ws_customers.append(["CustomerID", "Name", "Email"])
ws_customers.append(["C001", "John Doe", "john@example.com"])
ws_customers.append(["C002", "Jane Smith", "jane@example.com"])

ws_orders = wb.create_sheet("Orders")
ws_orders.append(["OrderID", "CustomerID", "OrderDate", "Total"])
ws_orders.append(["O001", "C001", "2024-01-15", 100.00])
ws_orders.append(["O002", "C002", "2024-01-16", 200.00])
ws_orders.append(["O003", "C001", "2024-01-17", 150.00])

wb.save(tmp_path)
print("✓ Created test file")

try:
    from rdfmap.generator.multisheet_analyzer import MultiSheetAnalyzer

    analyzer = MultiSheetAnalyzer(str(tmp_path))
    print(f"✓ Analyzer initialized with {len(analyzer.sheets)} sheets")

    # Check what we extracted
    for name, info in analyzer.sheets.items():
        print(f"\n  Sheet: {name}")
        print(f"    Rows: {info.row_count}")
        print(f"    Columns: {info.column_names}")
        print(f"    Identifiers: {info.identifier_columns}")
        print(f"    FK candidates: {info.foreign_key_candidates}")

    # Test detect_relationships
    print("\n=== Testing detect_relationships ===")
    relationships = analyzer.detect_relationships()

    print(f"\nFound {len(relationships)} relationship(s)")

    if relationships:
        for rel in relationships:
            print(f"  - {rel.source_sheet}.{rel.source_column} -> {rel.target_sheet}.{rel.target_column}")
            print(f"    Type: {rel.relationship_type}, Confidence: {rel.confidence:.2f}")

    # Check test assertion
    found_relationship = False
    for rel in relationships:
        if (rel.source_sheet == "Orders" and
            rel.target_sheet == "Customers" and
            "CustomerID" in rel.source_column and
            "CustomerID" in rel.target_column):
            found_relationship = True
            print(f"\n✓ Found expected relationship!")
            break

    if not found_relationship:
        print(f"\n❌ Expected relationship not found!")
        print(f"   Looking for: Orders.CustomerID -> Customers.CustomerID")
        sys.exit(1)

    print("\n✓ TEST PASSED")

except Exception as e:
    print(f"\n❌ ERROR: {e}")
    import traceback
    traceback.print_exc()
    sys.exit(1)
finally:
    if tmp_path.exists():
        tmp_path.unlink()
        print("✓ Cleaned up test file")


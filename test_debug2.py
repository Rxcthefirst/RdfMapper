#!/usr/bin/env python
"""Detailed debug to understand the issue."""

from pathlib import Path
from openpyxl import Workbook

# Create test file
tmp_path = Path("./test_debug2.xlsx")
wb = Workbook()
wb.remove(wb.active)

ws_customers = wb.create_sheet("Customers")
ws_customers.append(["CustomerID", "Name", "Email"])
ws_customers.append(["C001", "John Doe", "john@example.com"])
ws_customers.append(["C002", "Jane Smith", "jane@example.com"])

ws_orders = wb.create_sheet("Orders")
ws_orders.append(["OrderID", "CustomerID", "OrderDate", "Total"])
ws_orders.append(["O001", "C001", "2024-01-15", 100.00])
ws_orders.append(["O002", "C002", "2024-01-16", 200.00])
ws_orders.append(["O003", "C001", "2024-01-17", 150.00])

wb.save(tmp_path)

from rdfmap.generator.multisheet_analyzer import MultiSheetAnalyzer

analyzer = MultiSheetAnalyzer(str(tmp_path))

# Manual trace through detect_relationships
print("=== Manual trace of detect_relationships ===\n")

for source_name, source_sheet in analyzer.sheets.items():
    print(f"Processing sheet: {source_name}")
    print(f"  FK candidates: {source_sheet.foreign_key_candidates}")

    for fk_col, referenced_entity in source_sheet.foreign_key_candidates.items():
        print(f"\n  FK: {fk_col} -> entity: '{referenced_entity}'")

        # Test _find_matching_sheet
        target_sheet = analyzer._find_matching_sheet(referenced_entity)
        print(f"  _find_matching_sheet('{referenced_entity}') = {target_sheet}")

        if target_sheet:
            target_info = analyzer.sheets[target_sheet]
            print(f"  Target sheet columns: {target_info.column_names}")
            print(f"  Target identifier columns: {target_info.identifier_columns}")

            # Test _find_primary_key
            pk_col = analyzer._find_primary_key(target_info, referenced_entity)
            print(f"  _find_primary_key(..., '{referenced_entity}') = {pk_col}")

            if pk_col:
                # Test _analyze_relationship
                relationship = analyzer._analyze_relationship(
                    source_name, target_sheet,
                    fk_col, pk_col,
                    source_sheet, target_info
                )
                print(f"  _analyze_relationship result: {relationship}")
            else:
                print(f"  ❌ No primary key found")
        else:
            print(f"  ❌ No matching sheet found")
    print()

print("\n=== Calling detect_relationships ===")
relationships = analyzer.detect_relationships()
print(f"Result: {len(relationships)} relationships")

for i, rel in enumerate(relationships, 1):
    print(f"\n{i}. {rel.source_sheet}.{rel.source_column} -> {rel.target_sheet}.{rel.target_column}")
    print(f"   Type: {rel.relationship_type}, Confidence: {rel.confidence:.2f}")

tmp_path.unlink()
print(f"\n{'✓ SUCCESS' if len(relationships) > 0 else '❌ FAILED'}")


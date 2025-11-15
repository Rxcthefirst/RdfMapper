#!/usr/bin/env python
"""Trace through the entire detect_relationships flow."""
from pathlib import Path
from openpyxl import Workbook
from rdfmap.generator.multisheet_analyzer import MultiSheetAnalyzer

tmp_path = Path("./test_trace.xlsx")
wb = Workbook()
wb.remove(wb.active)

ws_customers = wb.create_sheet("Customers")
ws_customers.append(["CustomerID", "Name", "Email"])
ws_customers.append(["C001", "John Doe", "john@example.com"])
ws_customers.append(["C002", "Jane Smith", "jane@example.com"])

ws_orders = wb.create_sheet("Orders")
ws_orders.append(["OrderID", "CustomerID", "OrderDate", "Total"])
ws_orders.append(["O001", "C001", "2024-01-15", 100.00])
ws_orders.append(["O002", "C002", "2024-01-16", 200.00])
ws_orders.append(["O003", "C001", "2024-01-17", 150.00])

wb.save(tmp_path)

analyzer = MultiSheetAnalyzer(str(tmp_path))

print("="*60)
print("TRACING detect_relationships()")
print("="*60)

# Manually trace through detect_relationships
for source_name, source_sheet in analyzer.sheets.items():
    print(f"\n>>> Processing source sheet: {source_name}")
    print(f"    FK candidates: {source_sheet.foreign_key_candidates}")
    
    for fk_col, referenced_entity in source_sheet.foreign_key_candidates.items():
        print(f"\n    >>> FK: {fk_col} -> entity '{referenced_entity}'")
        
        # Step 1: Find matching sheet
        target_sheet = analyzer._find_matching_sheet(referenced_entity)
        print(f"        _find_matching_sheet('{referenced_entity}') = {target_sheet}")
        
        if target_sheet:
            target_info = analyzer.sheets[target_sheet]
            
            # Step 2: Find primary key
            pk_col = analyzer._find_primary_key(target_info, referenced_entity)
            print(f"        _find_primary_key(..., '{referenced_entity}') = {pk_col}")
            
            if pk_col:
                # Step 3: Analyze relationship
                print(f"        Calling _analyze_relationship(")
                print(f"            source_sheet={source_name},")
                print(f"            target_sheet={target_sheet},")
                print(f"            source_col={fk_col},")
                print(f"            target_col={pk_col},")
                print(f"            ...)")
                
                # Check sample data exists
                print(f"        source_sheet.sample_data is None: {source_sheet.sample_data is None}")
                print(f"        target_info.sample_data is None: {target_info.sample_data is None}")
                
                if source_sheet.sample_data is not None and target_info.sample_data is not None:
                    # Get values
                    source_values = source_sheet.sample_data[fk_col].drop_nulls()
                    target_values = target_info.sample_data[pk_col].drop_nulls()
                    
                    print(f"        Source values: {source_values.to_list()}")
                    print(f"        Target values: {target_values.to_list()}")
                    
                    source_set = set(source_values.to_list())
                    target_set = set(target_values.to_list())
                    
                    overlap = len(source_set & target_set)
                    overlap_ratio = overlap / len(source_set) if source_set else 0
                    
                    print(f"        Overlap: {overlap}/{len(source_set)} = {overlap_ratio:.2%}")
                    print(f"        Threshold: 30%")
                    
                    if overlap_ratio < 0.3:
                        print(f"        ❌ Overlap ratio too low, rejecting relationship")
                    else:
                        print(f"        ✓ Overlap ratio sufficient")
                
                relationship = analyzer._analyze_relationship(
                    source_name, target_sheet,
                    fk_col, pk_col,
                    source_sheet, target_info
                )
                
                print(f"        Result: {relationship}")
                
                if relationship:
                    print(f"        ✓ Relationship created!")
                else:
                    print(f"        ❌ _analyze_relationship returned None")
            else:
                print(f"        ❌ No primary key found")
        else:
            print(f"        ❌ No matching sheet found")

print("\n" + "="*60)
print("CALLING detect_relationships()")
print("="*60)

relationships = analyzer.detect_relationships()
print(f"\nResult: {len(relationships)} relationship(s)")

for rel in relationships:
    print(f"  - {rel.source_sheet}.{rel.source_column} -> {rel.target_sheet}.{rel.target_column}")

tmp_path.unlink()

if len(relationships) == 0:
    print("\n❌ TEST FAILED: No relationships detected")
else:
    print("\n✓ TEST PASSED")


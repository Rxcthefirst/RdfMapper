"""Data source parsers for CSV and XLSX files."""

from abc import ABC, abstractmethod
from pathlib import Path
from typing import Any, Dict, Generator, List, Optional

import pandas as pd
from openpyxl import load_workbook


class DataSourceParser(ABC):
    """Abstract base class for data source parsers."""

    @abstractmethod
    def parse(
        self, chunk_size: Optional[int] = None
    ) -> Generator[pd.DataFrame, None, None]:
        """Parse data source and yield DataFrames in chunks."""
        pass

    @abstractmethod
    def get_column_names(self) -> List[str]:
        """Get list of column names."""
        pass


class CSVParser(DataSourceParser):
    """Parser for CSV files."""

    def __init__(
        self,
        file_path: Path,
        delimiter: str = ",",
        has_header: bool = True,
        encoding: str = "utf-8",
    ):
        """Initialize CSV parser.
        
        Args:
            file_path: Path to CSV file
            delimiter: Column delimiter
            has_header: Whether first row is header
            encoding: File encoding
        """
        self.file_path = file_path
        self.delimiter = delimiter
        self.has_header = has_header
        self.encoding = encoding
        
        if not self.file_path.exists():
            raise FileNotFoundError(f"CSV file not found: {self.file_path}")

    def parse(
        self, chunk_size: Optional[int] = None
    ) -> Generator[pd.DataFrame, None, None]:
        """Parse CSV file and yield DataFrames.
        
        Args:
            chunk_size: Number of rows per chunk. If None, load entire file.
            
        Yields:
            DataFrames containing parsed data
        """
        header = 0 if self.has_header else None
        
        if chunk_size:
            # Stream large files in chunks
            for chunk in pd.read_csv(
                self.file_path,
                delimiter=self.delimiter,
                header=header,
                encoding=self.encoding,
                chunksize=chunk_size,
                keep_default_na=False,  # Preserve empty strings vs NaN
                na_values=[""],
            ):
                yield chunk
        else:
            # Load entire file
            df = pd.read_csv(
                self.file_path,
                delimiter=self.delimiter,
                header=header,
                encoding=self.encoding,
                keep_default_na=False,
                na_values=[""],
            )
            yield df

    def get_column_names(self) -> List[str]:
        """Get list of column names from CSV."""
        if not self.has_header:
            # Read first row to determine number of columns
            df_sample = pd.read_csv(
                self.file_path,
                delimiter=self.delimiter,
                nrows=1,
                header=None,
            )
            return [f"Column_{i}" for i in range(len(df_sample.columns))]
        
        # Read just the header
        df_sample = pd.read_csv(
            self.file_path,
            delimiter=self.delimiter,
            nrows=0,
        )
        return df_sample.columns.tolist()


class XLSXParser(DataSourceParser):
    """Parser for XLSX files."""

    def __init__(
        self,
        file_path: Path,
        sheet_name: Optional[str] = None,
        has_header: bool = True,
    ):
        """Initialize XLSX parser.
        
        Args:
            file_path: Path to XLSX file
            sheet_name: Name of sheet to read. If None, reads first sheet.
            has_header: Whether first row is header
        """
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.has_header = has_header
        
        if not self.file_path.exists():
            raise FileNotFoundError(f"XLSX file not found: {self.file_path}")

    def parse(
        self, chunk_size: Optional[int] = None
    ) -> Generator[pd.DataFrame, None, None]:
        """Parse XLSX file and yield DataFrames.
        
        Args:
            chunk_size: Number of rows per chunk. If None, load entire sheet.
            
        Yields:
            DataFrames containing parsed data
        """
        header = 0 if self.has_header else None
        
        if chunk_size:
            # For XLSX, we need to load and chunk manually
            # openpyxl doesn't support streaming reads efficiently
            df = pd.read_excel(
                self.file_path,
                sheet_name=self.sheet_name or 0,
                header=header,
                keep_default_na=False,
                na_values=[""],
            )
            
            # Yield in chunks
            for i in range(0, len(df), chunk_size):
                yield df.iloc[i : i + chunk_size]
        else:
            # Load entire sheet
            df = pd.read_excel(
                self.file_path,
                sheet_name=self.sheet_name or 0,
                header=header,
                keep_default_na=False,
                na_values=[""],
            )
            yield df

    def get_column_names(self) -> List[str]:
        """Get list of column names from XLSX."""
        if not self.has_header:
            # Read first row to determine columns
            df_sample = pd.read_excel(
                self.file_path,
                sheet_name=self.sheet_name or 0,
                nrows=1,
                header=None,
            )
            return [f"Column_{i}" for i in range(len(df_sample.columns))]
        
        # Read just the header
        df_sample = pd.read_excel(
            self.file_path,
            sheet_name=self.sheet_name or 0,
            nrows=0,
        )
        return df_sample.columns.tolist()

    def list_sheets(self) -> List[str]:
        """List all sheet names in the workbook."""
        wb = load_workbook(self.file_path, read_only=True)
        return wb.sheetnames


def create_parser(
    file_path: Path,
    delimiter: str = ",",
    has_header: bool = True,
    sheet_name: Optional[str] = None,
) -> DataSourceParser:
    """Factory function to create appropriate parser based on file extension.
    
    Args:
        file_path: Path to data file
        delimiter: CSV delimiter
        has_header: Whether file has header row
        sheet_name: XLSX sheet name (for Excel files)
        
    Returns:
        Appropriate parser instance
        
    Raises:
        ValueError: If file format is not supported
    """
    suffix = file_path.suffix.lower()
    
    if suffix == ".csv":
        return CSVParser(file_path, delimiter=delimiter, has_header=has_header)
    elif suffix in [".xlsx", ".xls"]:
        return XLSXParser(file_path, sheet_name=sheet_name, has_header=has_header)
    else:
        raise ValueError(f"Unsupported file format: {suffix}")

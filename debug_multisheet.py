#!/usr/bin/env python
"""Debug script to test multisheet analyzer."""

import sys
from pathlib import Path
from openpyxl import Workbook

# Create test file
tmp_path = Path("./test_multisheet_debug.xlsx")

wb = Workbook()
wb.remove(wb.active)

# Create Customers sheet
ws_customers = wb.create_sheet("Customers")
ws_customers.append(["CustomerID", "Name", "Email"])
ws_customers.append(["C001", "John Doe", "john@example.com"])
ws_customers.append(["C002", "Jane Smith", "jane@example.com"])

# Create Orders sheet
ws_orders = wb.create_sheet("Orders")
ws_orders.append(["OrderID", "CustomerID", "OrderDate", "Total"])
ws_orders.append(["O001", "C001", "2024-01-15", 100.00])
ws_orders.append(["O002", "C002", "2024-01-16", 200.00])
ws_orders.append(["O003", "C001", "2024-01-17", 150.00])

wb.save(tmp_path)
print(f"✓ Created test file: {tmp_path}")

# Import analyzer
from rdfmap.generator.multisheet_analyzer import MultiSheetAnalyzer

# Analyze
analyzer = MultiSheetAnalyzer(str(tmp_path))

print(f"\n✓ Analyzer initialized")
print(f"  Sheets found: {list(analyzer.sheets.keys())}")

for sheet_name, sheet_info in analyzer.sheets.items():
    print(f"\n  Sheet: {sheet_name}")
    print(f"    Rows: {sheet_info.row_count}")
    print(f"    Columns: {sheet_info.column_names}")
    print(f"    Identifier columns: {sheet_info.identifier_columns}")
    print(f"    FK candidates: {sheet_info.foreign_key_candidates}")
    if sheet_info.sample_data is not None:
        print(f"    Sample data shape: {sheet_info.sample_data.shape}")
        print(f"    Sample data:\n{sheet_info.sample_data}")

# Detect relationships
print("\n\n=== Detecting Relationships ===")
relationships = analyzer.detect_relationships()

print(f"\nRelationships found: {len(relationships)}")
for i, rel in enumerate(relationships, 1):
    print(f"\n  Relationship {i}:")
    print(f"    Source: {rel.source_sheet}.{rel.source_column}")
    print(f"    Target: {rel.target_sheet}.{rel.target_column}")
    print(f"    Type: {rel.relationship_type}")
    print(f"    Confidence: {rel.confidence:.2f}")
    if rel.cardinality_analysis:
        print(f"    Analysis: {rel.cardinality_analysis}")

# Cleanup
tmp_path.unlink()
print("\n✓ Cleaned up test file")

if len(relationships) == 0:
    print("\n❌ ERROR: No relationships detected!")
    sys.exit(1)
else:
    print(f"\n✓ SUCCESS: Found {len(relationships)} relationship(s)")


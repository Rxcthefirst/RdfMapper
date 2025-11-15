#!/usr/bin/env python
"""Minimal test of the matching logic."""

entity_name = "Customer"
entity_lower = entity_name.lower()

print(f"Entity: '{entity_name}' -> '{entity_lower}'")

# Generate plural forms
plural_forms = [
    entity_lower + 's',      # customer -> customers
    entity_lower + 'es',     # address -> addresses
    entity_lower + 'ies',    # category -> categories (handled separately)
]

if entity_lower.endswith('y') and len(entity_lower) > 1:
    plural_forms.append(entity_lower[:-1] + 'ies')

print(f"Plural forms: {plural_forms}")

# Test matching
sheets = {"Customers": None, "Orders": None}

for sheet_name in sheets.keys():
    sheet_lower = sheet_name.lower()
    print(f"\nTesting sheet '{sheet_name}' ('{sheet_lower}'):")
    print(f"  sheet_lower in plural_forms: {sheet_lower in plural_forms}")
    
    if sheet_lower in plural_forms:
        print(f"  ✓ MATCH!")

# Now test with actual analyzer
print("\n" + "="*50)
print("Testing with actual MultiSheetAnalyzer:")
print("="*50)

from pathlib import Path
from openpyxl import Workbook
from rdfmap.generator.multisheet_analyzer import MultiSheetAnalyzer

tmp_path = Path("./test_minimal.xlsx")
wb = Workbook()
wb.remove(wb.active)

ws_customers = wb.create_sheet("Customers")
ws_customers.append(["CustomerID", "Name"])
ws_customers.append(["C001", "John"])

ws_orders = wb.create_sheet("Orders")
ws_orders.append(["OrderID", "CustomerID"])
ws_orders.append(["O001", "C001"])

wb.save(tmp_path)

analyzer = MultiSheetAnalyzer(str(tmp_path))

# Test the method directly
result = analyzer._find_matching_sheet("Customer")
print(f"\n_find_matching_sheet('Customer') = {result}")

if result == "Customers":
    print("✓ SUCCESS: Method works correctly")
else:
    print(f"❌ FAILURE: Expected 'Customers', got {result}")

tmp_path.unlink()

